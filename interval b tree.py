import matplotlib.pyplot as plt
import numpy as np
import time
from openpyxl import load_workbook


class IntervalNode:
    def __init__(self):
        self.keys = []
        self.children = []

    def insert(self, key, data):
        pass


class IntervalLeafNode(IntervalNode):
    def __init__(self):
        super().__init__()
        self.data = []

    def insert(self, key, data):
        pass


class IntervalBTree:
    def __init__(self):
        self.root = None

    def insert(self, key, data):
        if self.root is None:
            self.root = IntervalLeafNode()
        self.root.insert(key, data)

    def search(self, key):
        if self.root is None:
            return None
        return self._search_helper(self.root, key)

    def _search_helper(self, node, key):
        if isinstance(node, IntervalLeafNode):
            for i, k in enumerate(node.keys):
                if k == key:
                    return node.data[i]
            return None
        else:
            for i, k in enumerate(node.keys):
                if key < k:
                    return self._search_helper(node.children[i], key)
            return self._search_helper(node.children[-1], key)


def read_excel_file(filename, column_name):
    data_dict = {}
    try:
        wb = load_workbook(filename)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        column_index = header.index(column_name)
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = row[column_index]
            data_dict.setdefault(key, []).append(row)
    except Exception as e:
        print("Error reading Excel file:", e)
    return data_dict


def main():
    start_time = time.time()
    interval_b_tree = IntervalBTree()
    column_name = 'Open'  # Change this to the desired column name
    data_dict = read_excel_file("C:\\Users\\sanka\\Desktop\\Reliance.xlsx", column_name)

    key_to_search = 1983  # Change this to the desired value from the specified column
    if key_to_search in data_dict:
        print("Rows corresponding to", column_name, "=", key_to_search, ":")
        for row in data_dict[key_to_search]:
            print(row)
    else:
        print("No rows found for", column_name, "=", key_to_search)

    end_time = time.time()

    elapsed_time = end_time - start_time
    print("Elapsed time:", elapsed_time, "seconds")


def generate_input_sizes(start, stop, step):
    return range(start, stop, step)


def measure_elapsed_time(input_sizes):
    elapsed_times = []
    for size in input_sizes:
        start_time = time.time()
        main()
        end_time = time.time()
        elapsed_time = end_time - start_time
        elapsed_times.append(elapsed_time)
    return elapsed_times


def plot_time_complexity(input_sizes, elapsed_times):
    plt.plot(input_sizes, elapsed_times, marker='o', linestyle='-')
    plt.title("Time Complexity")
    plt.xlabel("Input Size")
    plt.ylabel("Elapsed Time (seconds)")
    plt.grid(True)
    plt.show()


if __name__ == "__main__":
    start_input_size = 100
    stop_input_size = 1000
    step_input_size = 100

    input_sizes = generate_input_sizes(start_input_size, stop_input_size, step_input_size)
    elapsed_times = measure_elapsed_time(input_sizes)
    plot_time_complexity(input_sizes, elapsed_times)
