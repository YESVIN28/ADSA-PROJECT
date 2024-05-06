import matplotlib.pyplot as plt
import numpy as np
import time
from openpyxl import load_workbook


class IntervalNode:
    def insert(self, interval, data):
        pass

    def search(self, key):
        pass


class IntervalLeafNode(IntervalNode):
    def __init__(self, interval, data):
        self.intervals = [interval]
        self.data = [data]

    def insert(self, interval, data):
        self.intervals.append(interval)
        self.data.append(data)
        # Sorting intervals
        self.intervals, self.data = zip(*sorted(zip(self.intervals, self.data)))
        if len(self.intervals) > 3:
            return self.split()
        return None

    def split(self):
        mid = len(self.intervals) // 2
        new_leaf = IntervalLeafNode(self.intervals[mid:], self.data[mid:])
        self.intervals = self.intervals[:mid]
        self.data = self.data[:mid]
        return new_leaf

    def search(self, key):
        for i, interval in enumerate(self.intervals):
            if interval[0] <= key <= interval[1]:
                return self.data[i]
        return None


class IntervalInternalNode(IntervalNode):
    def __init__(self, interval, left_child, right_child):
        self.intervals = [interval]
        self.children = [left_child, right_child]

    def insert(self, interval, data):
        for i, inter in enumerate(self.intervals):
            if interval[0] <= inter[0]:
                if isinstance(self.children[i], IntervalLeafNode):
                    new_leaf = self.children[i].insert(interval, data)
                    if new_leaf:
                        self.intervals.insert(i, new_leaf.intervals[0])
                        self.children.insert(i + 1, new_leaf)
                else:
                    self.children[i].insert(interval, data)
                break
        else:
            if isinstance(self.children[-1], IntervalLeafNode):
                new_leaf = self.children[-1].insert(interval, data)
                if new_leaf:
                    self.intervals.append(new_leaf.intervals[0])
                    self.children.append(new_leaf)
            else:
                self.children[-1].insert(interval, data)

        if len(self.intervals) > 3:
            return self.split()
        return None

    def split(self):
        mid = len(self.intervals) // 2
        new_internal = IntervalInternalNode(self.intervals[mid], self, None)
        new_internal.children[1:] = self.children[mid + 1:]
        self.intervals = self.intervals[:mid]
        self.children = self.children[:mid + 1]
        return new_internal

    def search(self, key):
        for i, interval in enumerate(self.intervals):
            if key <= interval[1]:
                return self.children[i].search(key)
        return self.children[-1].search(key)


class IntervalBPlusTree:
    def __init__(self):
        self.root = None

    def insert(self, interval, data):
        if self.root is None:
            self.root = IntervalLeafNode(interval, data)
        else:
            new_root = self.root.insert(interval, data)
            if new_root:
                self.root = IntervalInternalNode(new_root.intervals[0], self.root, new_root)

    def search(self, key):
        if self.root is None:
            return None
        return self.root.search(key)


def read_excel_file(filename, column_name):
    data_dict = {}
    try:
        wb = load_workbook(filename)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        column_index = header.index(column_name)
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = row[column_index]
            data_dict.setdefault(key, []).append(row)
    except Exception as e:
        print("Error reading Excel file:", e)
    return data_dict


def main():
    start_time = time.time()
    interval_b_plus_tree = IntervalBPlusTree()
    column_name = 'Open'  # Change this to the desired column name
    data_dict = read_excel_file("C:\\Users\\sanka\\Desktop\\Reliance.xlsx", column_name)

    key_to_search = 475  # Change this to the desired value from the specified column
    if key_to_search in data_dict:
        print("Rows corresponding to", column_name, "=", key_to_search, ":")
        for row in data_dict[key_to_search]:
            print(row)
    else:
        print("No rows found for", column_name, "=", key_to_search)

    end_time_insert = time.time()

    # Measure the time taken for searching
    start_time_search = time.time()
    result = interval_b_plus_tree.search(key_to_search)
    end_time_search = time.time()

    if result:
        print("Data found:", result)
    else:
        print("Data not found.")

    elapsed_time_insert = end_time_insert - start_time
    elapsed_time_search = end_time_search - start_time_search
    print("Elapsed time (insertion):", elapsed_time_insert, "seconds")
    print("Elapsed time (searching):", elapsed_time_search, "seconds")

    return elapsed_time_insert, elapsed_time_search


def generate_input_sizes(start, stop, step):
    return range(start, stop, step)


def measure_elapsed_time(input_sizes):
    elapsed_times = []
    for size in input_sizes:
        start_time = time.time()
        main()
        end_time = time.time()
        elapsed_time = end_time - start_time
        elapsed_times.append(elapsed_time)
    return elapsed_times


def plot_time_complexity(input_sizes, insertion_times, search_times):
    plt.plot(input_sizes, insertion_times, marker='o', linestyle='-', label='Insertion Time')
    plt.plot(input_sizes, search_times, marker='s', linestyle='-', label='Search Time')
    plt.title("Time Complexity")
    plt.xlabel("Input Size")
    plt.ylabel("Elapsed Time (seconds)")
    plt.legend()
    plt.grid(True)
    plt.show()


if __name__ == "__main__":
    start_input_size = 100
    stop_input_size = 1000
    step_input_size = 100

    input_sizes = generate_input_sizes(start_input_size, stop_input_size, step_input_size)
    insertion_times = []
    search_times = []

    for size in input_sizes:
        insert_time, search_time = main()
        insertion_times.append(insert_time)
        search_times.append(search_time)

    plot_time_complexity(input_sizes, insertion_times, search_times)
