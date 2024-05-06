import matplotlib.pyplot as plt
import numpy as np
import time
from openpyxl import load_workbook


class BTreeNode:
    def __init__(self, leaf=False):
        self.keys = []
        self.children = []
        self.leaf = leaf

    def split_child(self, i, child):
        t = (degree - 1) // 2
        new_node = BTreeNode(leaf=child.leaf)
        self.children.insert(i + 1, new_node)
        self.keys.insert(i, child.keys[t])
        new_node.keys = child.keys[t + 1:]
        child.keys = child.keys[:t]

        if not child.leaf:
            new_node.children = child.children[t + 1:]
            child.children = child.children[:t + 1]

    def insert_non_full(self, key):
        i = len(self.keys) - 1
        if self.leaf:
            self.keys.append(0)
            while i >= 0 and key < self.keys[i]:
                self.keys[i + 1] = self.keys[i]
                i -= 1
            self.keys[i + 1] = key
        else:
            while i >= 0 and key < self.keys[i]:
                i -= 1
            i += 1
            if len(self.children[i].keys) == 2 * degree - 1:
                self.split_child(i, self.children[i])
                if key > self.keys[i]:
                    i += 1
            self.children[i].insert_non_full(key)

    def search(self, key):
        i = 0
        while i < len(self.keys) and key > self.keys[i]:
            i += 1
        if i < len(self.keys) and key == self.keys[i]:
            return self
        if self.leaf:
            return None
        return self.children[i].search(key)


class BTree:
    def __init__(self):
        self.root = None

    def insert(self, key):
        if self.root is None:
            self.root = BTreeNode(leaf=True)
            self.root.keys.append(key)
        else:
            if len(self.root.keys) == 2 * degree - 1:
                new_node = BTreeNode()
                new_node.children.append(self.root)
                new_node.split_child(0, self.root)
                i = 0
                if new_node.keys[0] < key:
                    i += 1
                new_node.children[i].insert_non_full(key)
                self.root = new_node
            else:
                self.root.insert_non_full(key)

    def search(self, key):
        return None if self.root is None else self.root.search(key)


def read_excel_file(filename, column_name):
    data_dict = {}
    try:
        wb = load_workbook(filename)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        column_index = header.index(column_name)
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = row[column_index]
            data_dict.setdefault(key, []).append(row)
    except Exception as e:
        print("Error reading Excel file:", e)
    return data_dict


def main():
    start_time = time.time()
    b_tree = BTree()
    column_name = 'Open'  # Change this to the desired column name
    data_dict = read_excel_file("C:\\Users\\sanka\\Desktop\\Reliance.xlsx", column_name)

    key_to_search = 475  # Change this to the desired value from the specified column
    if key_to_search in data_dict:
        print("Rows corresponding to", column_name, "=", key_to_search, ":")
        for row in data_dict[key_to_search]:
            print(row)
    else:
        print("No rows found for", column_name, "=", key_to_search)

    end_time_insert = time.time()

    # Measure the time taken for searching
    start_time_search = time.time()
    result = b_tree.search(key_to_search)
    end_time_search = time.time()

    if result:
        print("Data found:", result)
    else:
        print("Data not found.")

    elapsed_time_insert = end_time_insert - start_time
    elapsed_time_search = end_time_search - start_time_search
    print("Elapsed time (insertion):", elapsed_time_insert, "seconds")
    print("Elapsed time (searching):", elapsed_time_search, "seconds")

    return elapsed_time_insert, elapsed_time_search


def generate_input_sizes(start, stop, step):
    return range(start, stop, step)


def measure_elapsed_time(input_sizes):
    elapsed_times = []
    for size in input_sizes:
        start_time = time.time()
        main()
        end_time = time.time()
        elapsed_time = end_time - start_time
        elapsed_times.append(elapsed_time)
    return elapsed_times


def plot_time_complexity(input_sizes, insertion_times, search_times):
    plt.plot(input_sizes, insertion_times, marker='o', linestyle='-', label='Insertion Time')
    plt.plot(input_sizes, search_times, marker='s', linestyle='-', label='Search Time')
    plt.title("Time Complexity")
    plt.xlabel("Input Size")
    plt.ylabel("Elapsed Time (seconds)")
    plt.legend()
    plt.grid(True)
    plt.show()


if __name__ == "__main__":
    degree = 2  # Degree of B-tree

    start_input_size = 100
    stop_input_size = 1000
    step_input_size = 100

    input_sizes = generate_input_sizes(start_input_size, stop_input_size, step_input_size)
    insertion_times = []
    search_times = []

    for size in input_sizes:
        insert_time, search_time = main()
        insertion_times.append(insert_time)
        search_times.append(search_time)

    plot_time_complexity(input_sizes, insertion_times, search_times)
